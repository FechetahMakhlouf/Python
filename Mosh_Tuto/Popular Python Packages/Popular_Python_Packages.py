# ==============================
# 1. WHAT ARE APIs ?
# ==============================

# API = Application Programming Interface
#
# An API allows two applications
# to communicate with each other
#
# Example:
#
# Your Python application
#        ↓
#      API
#        ↓
# Another service/server


# ==============================
# REAL-LIFE EXAMPLES
# ==============================

# Weather App
#
# Your app requests weather data
# from a weather server API

# Payment Systems
#
# Websites use APIs like:
# - PayPal API
# - Stripe API
#
# to process payments

# Social Media APIs
#
# Platforms provide APIs:
# - Facebook API
# - Instagram API
# - Twitter/X API
#
# to access posts, messages, etc.


# ==============================
# HOW APIs WORK
# ==============================

# Usually:
#
# 1. Client sends a REQUEST
# 2. Server processes it
# 3. Server sends a RESPONSE

#
# Example:
#
# Python App
#     ↓ request
# API Server
#     ↓ response
# JSON Data


# ==============================
# HTTP METHODS
# ==============================

# APIs commonly use HTTP methods


# -------------------------
# GET
# -------------------------

# Used to GET data
#
# Example:
# Get weather information


# -------------------------
# POST
# -------------------------

# Used to SEND data
#
# Example:
# Create a new user


# -------------------------
# PUT
# -------------------------

# Used to UPDATE data
#
# Example:
# Update profile information


# -------------------------
# DELETE
# -------------------------

# Used to DELETE data
#
# Example:
# Delete a product


# ==============================
# API RESPONSE FORMATS
# ==============================

# APIs usually return:
#
# - JSON
# - XML
#
# Most modern APIs use JSON


# Example JSON response:
#
import numpy as np
from openpyxl.styles import Font
import openpyxl
import PyPDF2
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
from bs4 import BeautifulSoup
from twilio.rest import Client
from dotenv import load_dotenv
import os
import requests
{
    "name": "Makhlouf",
    "age": 22
}


# ==============================
# USING APIs IN PYTHON
# ==============================

# Python commonly uses:
# - requests
# - httpx
#
# to communicate with APIs


# Example using requests:


response = requests.get(
    "https://jsonplaceholder.typicode.com/users/1"
)

print(response.json())


# ==============================
# WHAT IS AN ENDPOINT ?
# ==============================

# An endpoint is a specific API URL
#
# Example:
#
# https://api.site.com/users
#
# Different endpoints perform
# different operations


# ==============================
# STATUS CODES
# ==============================

# APIs return HTTP status codes


# 200
# → Success

# 201
# → Created successfully

# 400
# → Bad request

# 401
# → Unauthorized

# 403
# → Forbidden

# 404
# → Resource not found

# 500
# → Server error


# ==============================
# API KEYS
# ==============================

# Many APIs require authentication
#
# Usually through:
# - API keys
# - tokens
#
# Example:
#
headers = {
    "Authorization": "Bearer TOKEN"
}


# ==============================
# WHY APIs ARE IMPORTANT
# ==============================

# APIs allow developers to:
# - use external services
# - connect applications
# - automate systems
# - build modern web/mobile apps


# ==============================
# POPULAR APIs
# ==============================

# - OpenAI API
# - Google Maps API
# - Stripe API
# - GitHub API
# - Weather APIs
# - Facebook Graph API

# ==============================
# 2. YELP API
# ==============================

# Yelp API allows developers to access:
# - restaurants
# - cafes
# - hotels
# - businesses
# - reviews
# - ratings
#
# using HTTP requests


# ==============================
# WHAT CAN YELP API DO ?
# ==============================

# With Yelp API you can:
#
# - search businesses
# - get restaurant details
# - get ratings/reviews
# - search by city/location
# - find nearby places


# ==============================
# STEP 1: CREATE A YELP ACCOUNT
# ==============================

# Create an account at:
#
# https://www.yelp.com/developers


# ==============================
# STEP 2: CREATE AN APP
# ==============================

# Create a new application
#
# Yelp provides:
# - API Key
#
# Example:
#
API_KEY = "your_api_key"


# ==============================
# STEP 3: INSTALL requests
# ==============================

# Install requests library
#
# Command:
# pip install requests


# ==============================
# STEP 4: SEND A REQUEST
# ==============================

# Import requests


# Your Yelp API key
API_KEY = "your_api_key"


# Yelp API endpoint
url = "https://api.yelp.com/v3/businesses/search"


# Request headers
#
# Authorization uses:
# Bearer TOKEN
headers = {
    "Authorization": f"Bearer {API_KEY}"
}


# Query parameters
#
# term:
# → what to search for
#
# location:
# → city/location
params = {
    "term": "restaurants",
    "location": "New York"
}


# Send GET request to Yelp API
response = requests.get(
    url,
    headers=headers,
    params=params
)


# ==============================
# STEP 5: CONVERT RESPONSE TO JSON
# ==============================

# Convert API response into Python dictionary
data = response.json()


# Print full JSON response
print(data)


# ==============================
# ACCESS BUSINESS DATA
# ==============================

# Loop through businesses
for business in data["businesses"]:

    # Print business name
    print(business["name"])

    # Print rating
    print("Rating:", business["rating"])

    # Print separator
    print("-" * 20)


# ==============================
# IMPORTANT JSON FIELDS
# ==============================

# Common fields returned:
#
business["name"]
# → business name
#
business["rating"]
# → rating score
#
business["price"]
# → price category
#
business["phone"]
# → phone number
#
business["location"]
# → address information


# ==============================
# STATUS CODES
# ==============================

# 200
# → Success

# 401
# → Invalid API key

# 404
# → Endpoint not found

# 500
# → Server error


# ==============================
# WHY YELP API IS USEFUL
# ==============================

# Useful for building:
# - restaurant finder apps
# - maps applications
# - local business search tools
# - recommendation systems


# ==============================
# EXAMPLE RESPONSE
# ==============================

# Example JSON:
#
{
    "businesses": [
        {
            "name": "Pizza House",
            "rating": 4.5
        }
    ]
}

# ==============================
# 3. HIDING API KEYS
# ==============================

# IMPORTANT:
#
# Never write API keys directly
# inside your source code
#
# BAD EXAMPLE:
#
# api_key = "123456SECRET"
#
# Because:
# - dangerous for GitHub
# - anyone can steal the key
# - security risk


# ==============================
# BEST SOLUTION:
# ENVIRONMENT VARIABLES
# ==============================

# Environment variables allow us
# to store secret information outside
# the source code


# ==============================
# STEP 1: INSTALL python-dotenv
# ==============================

# Command:
# pip install python-dotenv


# ==============================
# STEP 2: CREATE .env FILE
# ==============================

# Create a file named:
#
# .env
#
# Example content:
#
# API_KEY=your_secret_key_here


# IMPORTANT:
#
# Add .env to .gitignore
#
# so it is NOT uploaded to GitHub


# ==============================
# STEP 3: LOAD ENV VARIABLES
# ==============================

# Import os module
#
# Used to access environment variables

# Import load_dotenv


# Load variables from .env file
load_dotenv()


# Read API key from environment variables
api_key = os.getenv("API_KEY")


# Print the API key
# (for testing only)
print(api_key)


# ==============================
# USING THE API KEY
# ==============================

# Example with requests


headers = {
    "Authorization": f"Bearer {api_key}"
}


# Example request
response = requests.get(
    "https://api.example.com/data",
    headers=headers
)

print(response.status_code)


# ==============================
# PROJECT STRUCTURE
# ==============================

# Example:
#
# project/
# │
# ├── app.py
# ├── .env
# ├── .gitignore
# └── requirements.txt


# ==============================
# .gitignore EXAMPLE
# ==============================

# .env
# venv/
# __pycache__/


# ==============================
# WHY THIS IS IMPORTANT
# ==============================

# Protects:
# - API keys
# - passwords
# - tokens
# - database credentials


# ==============================
# COMMON ENV VARIABLES
# ==============================

# API_KEY
# SECRET_KEY
# DATABASE_URL
# ACCESS_TOKEN
# EMAIL_PASSWORD


# ==============================
# IMPORTANT
# ==============================

# Never upload:
# - .env
# - passwords
# - API keys
#
# to GitHub or public repositories

# ==============================
# 4. SENDING TEXT MESSAGES (TWILIO)
# ==============================

# Twilio is a cloud communication platform
#
# It allows developers to:
# - send SMS messages
# - make phone calls
# - send WhatsApp messages
# - build chat systems


# ==============================
# STEP 1: CREATE A TWILIO ACCOUNT
# ==============================

# Create an account at:
#
# https://www.twilio.com


# ==============================
# STEP 2: GET ACCOUNT INFORMATION
# ==============================

# From Twilio dashboard you will get:
#
# - Account SID
# - Auth Token
# - Twilio Phone Number


# ==============================
# STEP 3: INSTALL TWILIO PACKAGE
# ==============================

# Install Twilio library
#
# Command:
# pip install twilio


# ==============================
# STEP 4: IMPORT TWILIO CLIENT
# ==============================


# ==============================
# STEP 5: ACCOUNT CREDENTIALS
# ==============================

# Your Twilio Account SID
account_sid = "YOUR_ACCOUNT_SID"

# Your Twilio Auth Token
auth_token = "YOUR_AUTH_TOKEN"


# Create Twilio client object
client = Client(account_sid, auth_token)


# ==============================
# STEP 6: SEND SMS MESSAGE
# ==============================

# Send text message
message = client.messages.create(

    # Message content
    body="Hello from Python using Twilio!",

    # Twilio phone number
    from_="+123456789",

    # Receiver phone number
    to="+213000000000"
)


# ==============================
# STEP 7: PRINT MESSAGE SID
# ==============================

# Each sent message has a unique SID
print(message.sid)


# ==============================
# USING ENVIRONMENT VARIABLES
# ==============================

# Better and safer approach:
#
# Store credentials inside .env


# Example .env file:
#
# TWILIO_ACCOUNT_SID=xxxxxxxx
# TWILIO_AUTH_TOKEN=xxxxxxxx


# Example secure version:

# import os
# from dotenv import load_dotenv
#
# load_dotenv()
#
# account_sid = os.getenv("TWILIO_ACCOUNT_SID")
# auth_token = os.getenv("TWILIO_AUTH_TOKEN")


# ==============================
# POSSIBLE FEATURES
# ==============================

# Twilio supports:
#
# - SMS
# - WhatsApp
# - Voice calls
# - OTP verification
# - Chat bots
# - Notifications


# ==============================
# COMMON USE CASES
# ==============================

# - Login verification (OTP)
# - Appointment reminders
# - Order notifications
# - Emergency alerts
# - Marketing SMS


# ==============================
# IMPORTANT NOTES
# ==============================

# Trial accounts may:
# - require verified phone numbers
# - add "Sent from Twilio trial account"


# ==============================
# COMMON ERRORS
# ==============================

# Authentication Error
# → wrong SID or token
#
# Permission Error
# → unverified phone number
#
# Invalid Number
# → incorrect phone format


# ==============================
# PHONE NUMBER FORMAT
# ==============================

# Twilio uses international format
#
# Example:
# +213xxxxxxxxx
# +1xxxxxxxxxx

# ==============================
# 5. WEB SCRAPING
# ==============================

# Web scraping means:
# extracting data from websites automatically
#
# Python can:
# - download web pages
# - read HTML
# - extract information
#
# Common uses:
# - price tracking
# - news extraction
# - data collection
# - automation


# ==============================
# COMMON LIBRARIES
# ==============================

# requests
# → download web pages
#
# BeautifulSoup
# → parse HTML
#
# selenium
# → automate browsers


# ==============================
# STEP 1: INSTALL REQUIRED PACKAGES
# ==============================

# Install packages
#
# Command:
# pip install requests beautifulsoup4


# ==============================
# STEP 2: IMPORT LIBRARIES
# ==============================


# ==============================
# STEP 3: DOWNLOAD WEB PAGE
# ==============================

# Send GET request to website
response = requests.get(
    "https://example.com"
)

# Print status code
print(response.status_code)


# ==============================
# STEP 4: PARSE HTML
# ==============================

# Convert HTML into BeautifulSoup object
soup = BeautifulSoup(
    response.text,
    "html.parser"
)


# ==============================
# STEP 5: EXTRACT PAGE TITLE
# ==============================

# Get page title
print(soup.title)

# Get only title text
print(soup.title.text)


# ==============================
# STEP 6: FIND HTML ELEMENTS
# ==============================

# Find first paragraph
paragraph = soup.find("p")

print(paragraph.text)


# Find all links
links = soup.find_all("a")


# Loop through links
for link in links:

    # Get href attribute
    print(link.get("href"))


# ==============================
# EXTRACT BY CLASS
# ==============================

# Example HTML:
#
# <div class="product">Laptop</div>

# Find element by class name
# products = soup.find_all(
#     class_="product"
# )


# ==============================
# EXTRACT BY ID
# ==============================

# Example HTML:
#
# <h1 id="title">Hello</h1>

# title = soup.find(id="title")


# ==============================
# CSS SELECTORS
# ==============================

# BeautifulSoup supports CSS selectors

# Example:
# soup.select(".product")

# Example:
# soup.select("#title")


# ==============================
# SAVE SCRAPED DATA
# ==============================

# Example:
#
# with open("data.txt", "w") as file:
#     file.write(soup.text)


# ==============================
# COMMON WEB SCRAPING USES
# ==============================

# - News websites
# - E-commerce products
# - Job listings
# - Weather data
# - Social media analysis


# ==============================
# IMPORTANT NOTES
# ==============================

# Some websites block scraping
#
# Some websites require:
# - login
# - JavaScript rendering
#
# In such cases:
# selenium may be needed


# ==============================
# USING SELENIUM
# ==============================

# selenium controls a real browser
#
# Install:
# pip install selenium

# Example:
#
# from selenium import webdriver
#
# driver = webdriver.Chrome()
# driver.get("https://example.com")


# ==============================
# ETHICAL & LEGAL NOTES
# ==============================

# Always:
# - respect robots.txt
# - avoid excessive requests
# - follow website terms of service
#
# Some websites prohibit scraping


# ==============================
# EXAMPLE HTML
# ==============================

# Example page:
#
# <html>
#   <head>
#       <title>My Site</title>
#   </head>
#   <body>
#       <p>Hello World</p>
#   </body>
# </html>

# ==============================
# 6. BROWSER AUTOMATION
# (Mosh Hamedani Example)
# ==============================

# Browser automation means:
# controlling a web browser using Python
#
# Python can:
# - open websites
# - click buttons
# - fill forms
# - login automatically
# - scrape dynamic websites
#
# Most popular tool:
# selenium


# ==============================
# STEP 1: INSTALL SELENIUM
# ==============================

# Command:
# pip install selenium


# ==============================
# STEP 2: INSTALL A WEB DRIVER
# ==============================

# Selenium needs a browser driver
#
# Example:
# ChromeDriver for Google Chrome
#
# Download from:
# https://chromedriver.chromium.org
#
# IMPORTANT:
# ChromeDriver version must match
# your Chrome browser version


# ==============================
# STEP 3: IMPORT SELENIUM
# ==============================


# ==============================
# STEP 4: OPEN THE BROWSER
# ==============================

# Create Chrome browser object
browser = webdriver.Chrome()


# ==============================
# STEP 5: OPEN A WEBSITE
# ==============================

# Open Python website
browser.get("https://github.com")


# ==============================
# STEP 6: FIND ELEMENTS
# ==============================

# Find search input by name
search_box = browser.find_element(
    By.NAME,
    "q"
)


# ==============================
# STEP 7: TYPE INSIDE INPUT
# ==============================

# Type text into search box
search_box.send_keys("python")


# Press ENTER key
search_box.send_keys(Keys.RETURN)


# Wait a few seconds
time.sleep(3)


# ==============================
# STEP 8: EXTRACT INFORMATION
# ==============================

# Find repository links
repositories = browser.find_elements(
    By.CSS_SELECTOR,
    "a"
)

# Print first few links
for repo in repositories[:5]:

    # Print visible text
    print(repo.text)


# ==============================
# STEP 9: CLOSE THE BROWSER
# ==============================

browser.quit()


# ==============================
# OTHER WAYS TO FIND ELEMENTS
# ==============================

# By.ID
# browser.find_element(By.ID, "username")

# By.CLASS_NAME
# browser.find_element(By.CLASS_NAME, "btn")

# By.TAG_NAME
# browser.find_element(By.TAG_NAME, "input")

# By.CSS_SELECTOR
# browser.find_element(By.CSS_SELECTOR, ".btn")

# By.XPATH
# browser.find_element(By.XPATH, "//button")


# ==============================
# AUTOMATING LOGIN
# ==============================

# Example:
#
# username = browser.find_element(By.ID, "username")
# password = browser.find_element(By.ID, "password")
#
# username.send_keys("admin")
# password.send_keys("1234")
#
# password.send_keys(Keys.RETURN)


# ==============================
# HEADLESS MODE
# ==============================

# Headless mode runs browser
# WITHOUT opening a visible window

# Example:
#
# from selenium.webdriver.chrome.options import Options
#
# options = Options()
# options.add_argument("--headless")
#
# browser = webdriver.Chrome(options=options)


# ==============================
# COMMON USE CASES
# ==============================

# - Automated testing
# - Form submission
# - Web scraping
# - Social media automation
# - Bots
# - Data extraction


# ==============================
# IMPORTANT NOTES
# ==============================

# Some websites:
# - detect automation
# - block bots
#
# Browser automation may be slower
# than normal API requests


# ==============================
# SELENIUM VS REQUESTS
# ==============================

# requests:
# - faster
# - no browser
# - static HTML only
#
# selenium:
# - real browser
# - supports JavaScript
# - slower but more powerful

# ==============================
# 7. WORKING WITH PDFS (PyPDF2)
# ==============================

# PyPDF2 is a Python library used to:
# - read PDF files
# - extract text
# - merge PDFs
# - split PDFs
# - rotate pages
# - encrypt PDFs


# ==============================
# STEP 1: INSTALL PyPDF2
# ==============================

# Command:
# pip install PyPDF2


# ==============================
# STEP 2: IMPORT PyPDF2
# ==============================


# ==============================
# READING A PDF FILE
# ==============================

# Open PDF file in binary read mode
with open("document.pdf", "rb") as file:

    # Create PDF reader object
    reader = PyPDF2.PdfReader(file)

    # Print number of pages
    print("Pages:", len(reader.pages))

    # Get first page
    page = reader.pages[0]

    # Extract text from page
    text = page.extract_text()

    # Print extracted text
    print(text)


# ==============================
# LOOP THROUGH ALL PAGES
# ==============================

with open("document.pdf", "rb") as file:

    reader = PyPDF2.PdfReader(file)

    # Loop through all pages
    for page in reader.pages:

        # Extract text
        text = page.extract_text()

        print(text)


# ==============================
# MERGING PDF FILES
# ==============================

# Create PDF merger object
merger = PyPDF2.PdfMerger()

# Add PDF files
merger.append("file1.pdf")
merger.append("file2.pdf")

# Save merged PDF
merger.write("merged.pdf")

# Close merger
merger.close()


# ==============================
# SPLITTING PDF PAGES
# ==============================

with open("document.pdf", "rb") as file:

    reader = PyPDF2.PdfReader(file)

    # Create writer object
    writer = PyPDF2.PdfWriter()

    # Add first page only
    writer.add_page(reader.pages[0])

    # Save new PDF
    with open("first_page.pdf", "wb") as output:
        writer.write(output)


# ==============================
# ROTATING A PAGE
# ==============================

with open("document.pdf", "rb") as file:

    reader = PyPDF2.PdfReader(file)

    writer = PyPDF2.PdfWriter()

    # Get first page
    page = reader.pages[0]

    # Rotate page 90 degrees
    page.rotate(90)

    # Add rotated page
    writer.add_page(page)

    # Save rotated PDF
    with open("rotated.pdf", "wb") as output:
        writer.write(output)


# ==============================
# ENCRYPTING A PDF
# ==============================

with open("document.pdf", "rb") as file:

    reader = PyPDF2.PdfReader(file)

    writer = PyPDF2.PdfWriter()

    # Add all pages
    for page in reader.pages:
        writer.add_page(page)

    # Add password protection
    writer.encrypt("1234")

    # Save encrypted PDF
    with open("encrypted.pdf", "wb") as output:
        writer.write(output)


# ==============================
# DECRYPTING A PDF
# ==============================

with open("encrypted.pdf", "rb") as file:

    reader = PyPDF2.PdfReader(file)

    # Decrypt using password
    reader.decrypt("1234")

    # Extract text
    text = reader.pages[0].extract_text()

    print(text)


# ==============================
# COMMON USE CASES
# ==============================

# - PDF text extraction
# - PDF merging
# - Report generation
# - Document automation
# - PDF protection
# - Splitting large PDFs


# ==============================
# IMPORTANT NOTES
# ==============================

# PyPDF2 works best with:
# - text-based PDFs
#
# It may NOT work well with:
# - scanned PDFs
# - image-only PDFs

# For scanned PDFs:
# OCR tools may be needed
#
# Example:
# pytesseract


# ==============================
# OTHER PDF LIBRARIES
# ==============================

# reportlab
# → create PDFs
#
# pdfplumber
# → advanced text extraction
#
# fitz / pymupdf
# → fast PDF processing
#
# pdfminer
# → detailed PDF parsing

# ==============================
# 8. WORKING WITH EXCEL SPREADSHEETS
# ==============================

# Python can work with Excel files:
# - read spreadsheets
# - write spreadsheets
# - modify cells
# - create reports
# - automate Excel tasks
#
# Most popular library:
# openpyxl


# ==============================
# STEP 1: INSTALL openpyxl
# ==============================

# Command:
# pip install openpyxl


# ==============================
# STEP 2: IMPORT openpyxl
# ==============================


# ==============================
# CREATING A NEW EXCEL FILE
# ==============================

# Create workbook object
workbook = openpyxl.Workbook()

# Get active worksheet
sheet = workbook.active

# Rename worksheet
sheet.title = "Students"


# ==============================
# WRITING DATA TO CELLS
# ==============================

# Write values into cells
sheet["A1"] = "Name"
sheet["B1"] = "Age"

sheet["A2"] = "Makhlouf"
sheet["B2"] = 22

sheet["A3"] = "Ahmed"
sheet["B3"] = 25


# ==============================
# SAVE EXCEL FILE
# ==============================

# Save workbook
workbook.save("students.xlsx")


# ==============================
# LOADING AN EXISTING FILE
# ==============================

# Load workbook
workbook = openpyxl.load_workbook(
    "students.xlsx"
)

# Select worksheet
sheet = workbook["Students"]


# ==============================
# READING CELL VALUES
# ==============================

# Read single cell
print(sheet["A2"].value)

# Read another cell
print(sheet["B2"].value)


# ==============================
# LOOP THROUGH ROWS
# ==============================

# Iterate through rows
for row in sheet.iter_rows(
    min_row=2,
    values_only=True
):

    print(row)


# ==============================
# LOOP THROUGH COLUMNS
# ==============================

# Iterate through columns
for column in sheet.iter_cols(
    min_col=1,
    max_col=2,
    values_only=True
):

    print(column)


# ==============================
# APPEND NEW ROWS
# ==============================

# Add new row
sheet.append(["Yacine", 30])

# Save changes
workbook.save("students.xlsx")


# ==============================
# GET SHEET DIMENSIONS
# ==============================

# Number of rows
print(sheet.max_row)

# Number of columns
print(sheet.max_column)


# ==============================
# ACCESS ROWS/COLUMNS
# ==============================

# Access first row
print(sheet[1])

# Access column A
print(sheet["A"])


# ==============================
# FORMATTING CELLS
# ==============================


# Make header bold
sheet["A1"].font = Font(bold=True)
sheet["B1"].font = Font(bold=True)

# Save changes
workbook.save("students.xlsx")


# ==============================
# CREATING MULTIPLE SHEETS
# ==============================

# Create new worksheet
workbook.create_sheet("Teachers")

# Save workbook
workbook.save("students.xlsx")


# ==============================
# DELETE A SHEET
# ==============================

# Remove worksheet
# del workbook["Teachers"]

# workbook.save("students.xlsx")


# ==============================
# COMMON USE CASES
# ==============================

# - Reports
# - Financial spreadsheets
# - Student management
# - Automation
# - Data analysis
# - Exporting data


# ==============================
# IMPORTANT NOTES
# ==============================

# openpyxl supports:
# - .xlsx files
#
# It does NOT fully support:
# - old .xls files


# ==============================
# OTHER EXCEL LIBRARIES
# ==============================

# pandas
# → data analysis
#
# xlsxwriter
# → advanced Excel writing
#
# xlrd
# → older Excel file reading


# ==============================
# EXAMPLE FILE STRUCTURE
# ==============================

# students.xlsx
#
# ┌──────────┬─────┐
# │ Name     │ Age │
# ├──────────┼─────┤
# │ Makhlouf │ 22  │
# │ Ahmed    │ 25  │
# └──────────┴─────┘

# =========================================================
# 9- Command Query Separation Principle (CQS)
# =========================================================
#
# CQS means:
#
# 1) A COMMAND:
#    - Changes data/state
#    - Does NOT return important information
#
# 2) A QUERY:
#    - Returns information
#    - Does NOT change data/state
#
# In this example with Excel:
#
# QUERY examples:
# - reading cell values
# - printing data
#
# COMMAND examples:
# - appending new rows
# - saving the workbook
#
# =========================================================


# ---------------------------------------------------------
# QUERY:
# Load the workbook (Excel file)
# ---------------------------------------------------------
# load_workbook() opens the Excel file so we can read/write.
wb = openpyxl.load_workbook("transactions.xlsx")

# ---------------------------------------------------------
# QUERY:
# Access a worksheet
# ---------------------------------------------------------
# Here we select the sheet named "Sheet1".
sheet = wb["Sheet1"]

# =========================================================
# QUERY SECTION
# =========================================================
#
# This section ONLY reads data.
# It does NOT modify the Excel file.
#
# =========================================================

# Loop through rows from 1 to 9
# range(1, 10) means:
# start at 1
# stop before 10
for row in range(1, 10):

    # -----------------------------------------------------
    # QUERY:
    # Read a specific cell
    # -----------------------------------------------------
    #
    # sheet.cell(row, column)
    #
    # row     -> current row number
    # column  -> 1 means column A
    #
    # Examples:
    # cell(1,1) = A1
    # cell(2,1) = A2
    # cell(3,1) = A3
    #
    cell = sheet.cell(row, 1)

    # -----------------------------------------------------
    # QUERY:
    # Get the value inside the cell
    # -----------------------------------------------------
    #
    # .value reads the content stored in the cell.
    #
    print(cell.value)

# =========================================================
# COMMAND SECTION
# =========================================================
#
# This section MODIFIES the workbook.
#
# =========================================================

# ---------------------------------------------------------
# COMMAND:
# Add a new row to the sheet
# ---------------------------------------------------------
#
# append() changes the Excel file in memory.
#
# This adds:
# A column -> 1
# B column -> 2
# C column -> 3
#
sheet.append([1, 2, 3])

# ---------------------------------------------------------
# COMMAND:
# Save changes into a new Excel file
# ---------------------------------------------------------
#
# save() writes modifications to disk.
#
# A new file named "transactions2.xlsx" will be created.
#
wb.save("transactions2.xlsx")

# =========================================================
# SUMMARY
# =========================================================
#
# QUERIES (read only):
# - load_workbook()
# - accessing sheet
# - cell.value
# - print()
#
# COMMANDS (modify data):
# - append()
# - save()
#
# The idea of CQS:
#
# Separate:
#   Reading data  (Queries)
# from:
#   Modifying data (Commands)
#
# This makes code:
# - cleaner
# - easier to debug
# - easier to maintain
#
# =========================================================

# =========================================================
# NUMPY WITH PYTHON (Code With Mosh)
# =========================================================
#
# NumPy = Numerical Python
#
# It is a powerful library used for:
# - arrays
# - mathematics
# - statistics
# - machine learning
# - data science
#
# Main advantage:
# NumPy arrays are MUCH faster than Python lists.
#
# =========================================================

# ---------------------------------------------------------
# IMPORT NUMPY
# ---------------------------------------------------------


# =========================================================
# CREATING ARRAYS
# =========================================================

# ---------------------------------------------------------
# Create a simple array
# ---------------------------------------------------------

numbers = np.array([1, 2, 3])

print(numbers)

# Output:
# [1 2 3]

# ---------------------------------------------------------
# Array with multiple rows (2D array)
# ---------------------------------------------------------

matrix = np.array([
    [1, 2, 3],
    [4, 5, 6]
])

print(matrix)

# Output:
# [[1 2 3]
#  [4 5 6]]

# =========================================================
# ACCESSING ELEMENTS
# =========================================================

# ---------------------------------------------------------
# Access row and column
# ---------------------------------------------------------
#
# Syntax:
# array[row][column]
#
# Row index starts from 0
# Column index starts from 0
#

print(matrix[0][1])

# Output:
# 2
#
# Explanation:
# row 0 = [1, 2, 3]
# column 1 = 2

# =========================================================
# ARRAY SHAPE
# =========================================================

# ---------------------------------------------------------
# shape gives:
# (rows, columns)
# ---------------------------------------------------------

print(matrix.shape)

# Output:
# (2, 3)

# Means:
# 2 rows
# 3 columns

# =========================================================
# CREATING SPECIAL ARRAYS
# =========================================================

# ---------------------------------------------------------
# Array filled with zeros
# ---------------------------------------------------------

zeros = np.zeros((3, 4))

print(zeros)

# Creates:
# 3 rows
# 4 columns
# all values = 0

# ---------------------------------------------------------
# Array filled with ones
# ---------------------------------------------------------

ones = np.ones((2, 3))

print(ones)

# ---------------------------------------------------------
# Create a sequence of numbers
# ---------------------------------------------------------
#
# arange(start, stop)
# stop is NOT included
#

numbers = np.arange(1, 10)

print(numbers)

# Output:
# [1 2 3 4 5 6 7 8 9]

# ---------------------------------------------------------
# Create decimal sequence
# ---------------------------------------------------------
#
# linspace(start, stop, count)
#

line = np.linspace(0, 5, 10)

print(line)

# Creates 10 numbers between 0 and 5

# =========================================================
# BASIC OPERATIONS
# =========================================================

array1 = np.array([1, 2, 3])
array2 = np.array([10, 20, 30])

# ---------------------------------------------------------
# Addition
# ---------------------------------------------------------

print(array1 + array2)

# Output:
# [11 22 33]

# ---------------------------------------------------------
# Multiplication
# ---------------------------------------------------------

print(array1 * array2)

# Output:
# [10 40 90]

# ---------------------------------------------------------
# Add a single number to all elements
# ---------------------------------------------------------

print(array1 + 5)

# Output:
# [6 7 8]

# =========================================================
# MATHEMATICAL FUNCTIONS
# =========================================================

numbers = np.array([1, 2, 3, 4])

# ---------------------------------------------------------
# Sum
# ---------------------------------------------------------

print(numbers.sum())

# Output:
# 10

# ---------------------------------------------------------
# Mean (average)
# ---------------------------------------------------------

print(numbers.mean())

# Output:
# 2.5

# ---------------------------------------------------------
# Maximum value
# ---------------------------------------------------------

print(numbers.max())

# Output:
# 4

# ---------------------------------------------------------
# Minimum value
# ---------------------------------------------------------

print(numbers.min())

# Output:
# 1

# =========================================================
# RANDOM VALUES
# =========================================================

# ---------------------------------------------------------
# Random decimal numbers between 0 and 1
# ---------------------------------------------------------

random_numbers = np.random.random((2, 3))

print(random_numbers)

# Creates:
# 2 rows
# 3 columns
# with random values

# ---------------------------------------------------------
# Random integers
# ---------------------------------------------------------
#
# randint(start, stop, size)
#

random_ints = np.random.randint(1, 10, (2, 2))

print(random_ints)

# Creates random integers from 1 to 9

# =========================================================
# RESHAPING ARRAYS
# =========================================================

numbers = np.array([1, 2, 3, 4, 5, 6])

# ---------------------------------------------------------
# reshape(rows, columns)
# ---------------------------------------------------------

reshaped = numbers.reshape(2, 3)

print(reshaped)

# Output:
# [[1 2 3]
#  [4 5 6]]

# =========================================================
# COPYING ARRAYS
# =========================================================

original = np.array([1, 2, 3])

# ---------------------------------------------------------
# copy()
# ---------------------------------------------------------
#
# Without copy():
# modifying one array may affect the other.
#

copied = original.copy()

copied[0] = 100

print(original)

# Output:
# [1 2 3]

print(copied)

# Output:
# [100   2   3]

# =========================================================
# SUMMARY
# =========================================================
#
# Important NumPy concepts:
#
# np.array()      -> create arrays
# .shape          -> array dimensions
# np.zeros()      -> array of zeros
# np.ones()       -> array of ones
# np.arange()     -> sequence of integers
# np.linspace()   -> sequence with equal spacing
# reshape()       -> change dimensions
# sum()           -> total
# mean()          -> average
# max()           -> largest value
# min()           -> smallest value
#
# NumPy is heavily used in:
# - AI
# - Machine Learning
# - Data Science
# - Scientific Computing
#
# =========================================================
